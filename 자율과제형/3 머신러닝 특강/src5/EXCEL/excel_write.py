# 모듈 로딩 -------------------------------------------
import openpyxl

# 데이터 변수 선언 -------------------------------------
filename = "../DATA/stats_100701.xlsx"
B_CODE = 66

# 엑셀 특정 데이터 추출 --------------------------------
book = openpyxl.load_workbook(filename)
sheet = book.active
MAX_COL = sheet.max_column
MAX_ROW = sheet.max_row

# 서울 제외 인구 추출 파일 저장 ---------------------------------
for i in range(0, MAX_COL-1):
    total = int(sheet[str(chr(i + B_CODE)) + "4"].value)   # 전체
    seoul = int(sheet[str(chr(i + B_CODE)) + "5"].value)   # 서울
    output = total - seoul
    print("서울 제외 인구 =", i, output)

    # 새로운 ROW 추가
    sheet[str(chr(i + B_CODE)) + str(MAX_ROW+1)] = output
    cell = sheet[str(chr(i + B_CODE)) + str(MAX_ROW+1)]
    cell.font = openpyxl.styles.Font(size=14,color="FF0000")
    cell.number_format = cell.number_format

# 엑셀 파일 저장
filename = "../DATA/population.xlsx"
book.save(filename)
print("ok")
