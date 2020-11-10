# 모듈로딩 -------------------------------------------------
from openpyxl import Workbook

# 데이터 변수 선언 -------------------------------------------
filename = '../DATA/sample02.xlsx'
names=['Name','Tom', 'Jane', 'Park', 'Kim']
ages=['Age', 14, 21, 19, 20]

# 엑셀 파일 생성 --------------------------------------------
wb = Workbook()                     # 엑셀 파일 생성, Sheet1 자동 생성
ws = wb.active                      # 시트 활성화
ws.title = 'new sheet'              # 시트명 변경
ws['A1'] = 'Language'               # 시트 데이터 삽입
ws['B1'] = 'Create'

#wb.remove_sheet(ws)                 # 1번째 시트 삭제
ws2=wb.create_sheet('data_sheet')    # 2번째 시트 생성
"""
ws2.cell(1,1).value='Name'
ws2.cell(1,2).value='Age'
ws2.cell(2,1).value='Tom'
ws2.cell(2,2).value='010-111-2222'
"""
for r in range(len(names)):
    print(r)
    ws2.cell(r+1, 1).value = names[r]

for r in range(len(ages)):
    ws2.cell(r+1,2).value = ages[r]

wb.active=ws2                        # 활성화 시트 설정
wb.save(filename = filename)         # 엑셀 파일 생성