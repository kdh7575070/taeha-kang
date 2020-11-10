# 모듈 로딩 ------------------------------------------------
import openpyxl

# 데이터 변수 선언 ------------------------------------------
filename = "../DATA/sample.xlsx"

# 엑셀 파일 열기 ---------------------------------------------
book = openpyxl.load_workbook(filename)

# 시트 추출
sheet = book.worksheets[0]
print('sheet =>',  sheet.title)
print('sheet.max_row = {}, sheet.max_column={}'.format(sheet.max_row, sheet.max_column))

rowdata=[]
for row in sheet.rows:
    rowdata.append([
        row[0].value,
        row[sheet.max_column-1].value
    ])

print('len(rowdata) => ', len(rowdata), rowdata)

"""

# 시트의 각 행을 순서대로 추출하기 --- (※3)
DATA = []
skip = 1
print('max_column={}, max_row={}'.format(sheet.max_column ,sheet.max_row) )

for row in sheet.rows:
    if skip > 4:
        DATA.append([
            row[0].value,
            row[sheet.max_column-1].value
        ])
        print('row[0].value={}, max_row={} {}'.format(row[0].value ,row[sheet.max_column-1].value, type(row[sheet.max_column-1].value)))
    else:
        skip +=1

print('len(DATA) => ', len(DATA))

for i in DATA:
    print(i)
# 데이터를 인구 순서로 정렬합니다.
DATA = sorted(DATA, key=lambda x:x[1])

# 하위 5위를 출력합니다.
for i, a in enumerate(DATA):
    if (i >= 5): break
    print(i+1, a[0], int(a[1]))
"""