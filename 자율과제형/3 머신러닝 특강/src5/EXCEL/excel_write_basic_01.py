# 모듈로딩 -------------------------------------------------
from openpyxl import Workbook

# 데이터 변수 선언 -------------------------------------------
filename = '../DATA/sample.xlsx'

# 엑셀 파일 생성 --------------------------------------------
wb = Workbook()                     # 엑셀 파일 생성, Sheet1 자동 생성
ws = wb.active                      # 시트 활성화
ws.title = 'new sheet'              # 시트명 변경
ws['A1'] = 'Language'               # 시트 데이터 삽입
ws['B1'] = 'Create'

wb.save(filename = filename)